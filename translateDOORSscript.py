# ParseScript in project Testparser
# This is the latest file which has been sanitised for errors and warnings - 28-03-19

"""
 A. Moghul - April 2019

 This takes a test script file and matches against a set of rules to convert the script file into
 a structured file format (defined by Labview)

 The input file MUST be a Labview pre-processed file - This takes the Inspect commands
 and puts them within the Script as well as other things like removing the numbering.  See Raguprasad Rao


 The parsing relies on determining which CDNU (CDNU1/CDNU2 or Both CDNUS) the tests will be carried out on.
 As the test scripts are non consistent, the following rules will apply:
 Where CDNU is explicitly noted (CDNU1/2/Both), the actions will be performed appropriately
 Where CDNU has not been identified, the actions will default to CDNU1

 The Gui chosen is TKinter, purely because I needed a quick way to design a form. The code for this has been kept
 modular, so if you needed to change the GUI like Qt or Kivy, it should be very easy

 Disclaimer: This is my first Python program, so probably there are hundreds of better ways to do this! :)

"""

import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
# from word2number import w2n
import logging
import sys
import getopt
from tkinter import filedialog
from tkinter import scrolledtext
from tkinter import *



# GLOBAL Definitions

SEPCH = ":"
COMMENT = "### "
INPUT_COL = 2
PROCEDURE_COL = 2
CDNU_COL = 3
OUTPUT_COL = 4
ERROR_COL = 5

def open_excel(xl_filename: str) -> Workbook:
    """
    Opens an Excel file for processing, given the pathname
    :param xl_filename:
    :return:  Workbook
    """
    logging.info("open_excel")

    try:
        wbook = load_workbook(xl_filename)  # Load file
        logging.info(f"Opened file {xl_filename}")
        return wbook
    except FileNotFoundError:
        print(f"File {xl_filename} not found, Exiting...")
        exit(-1)
    except OSError:
        print(f"{xl_filename} Error Saving file... Exiting...")
        exit(-1)
    except PermissionError:
        print(f"Error reading file {xl_filename} ... Exiting...")
        exit(-1)
    except Exception as ex:
        print(f'Error when opening {xl_filename} :', (str(ex)))
        exit(-1)


def close_excel(wbook: Workbook, xl_filename: str):
    """
    Closes the excel file

    :param wbook:
    :param xl_filename:
    :return:
    """
    logging.info("close_excel")

    try:
        wbook.save(xl_filename)
        wbook.close()
        logging.info(f"Closed file {xl_filename}")
    except FileNotFoundError:
        print(f"File {xl_filename} not found when Closing, Exiting...")
        exit(-1)
    except OSError:
        print(f"{xl_filename} Error Saving file... Exiting...")
        exit(-1)
    except PermissionError:
        print(f"{xl_filename} Error Saving file... Exiting...")
        exit(-1)
    except Exception as ex:
        print(f'Error when Closing {xl_filename} :', (str(ex)))
        exit(-1)


def process_arinc(cell_val, cell, work_sheet):

    """
        Processes the main ARINC keywords
        Any commands which set multiple values have been ignored.
    """

    global SEPCH
    global COMMENT
    global INPUT_COL
    global CDNU_COL
    global OUTPUT_COL
    global ERROR_COL

    rec_arinc = re.compile("ARINC Simulator\s?:?\s?(?P<Set>set)\s?(?P<Item>.*)to(?P<Val>.*)(?P<Bracket> \(.*\))",
                           re.IGNORECASE)

    a1 = rec_arinc.search(cell_val)

    if a1:
        try:
            set_item = a1.group('Set')
            item = a1.group('Item')
            val = a1.group('Val')
            bracket = a1.group('Bracket')
        except (NameError, AttributeError):
            logging.debug(f"{cell.row} Unable to determine sub matches for ARINC {cell_val}")
            pass
        else:
            constructed_str = "ARINC" + SEPCH + \
                              "SET" + SEPCH + \
                              str(set_item) + SEPCH + \
                              str(item).upper() + \
                              str(val) + SEPCH + \
                              COMMENT + SEPCH + \
                              str(bracket)

            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str

            # Too many variations of the value for not enough gain - do this manually - but alert the user
            work_sheet.cell(row=cell.row, column=ERROR_COL).value = "ALERT!"

            logging.debug(f"ARINC = {constructed_str}")


def process_waitfor(cell_val, cell, work_sheet):

    """
        Processes the wait commands. Translates some wait commands from a string to number
        e.g wait 31 seconds becomes 31
        Any commands which set multiple values have been ignored.
    """

    global SEPCH
    global COMMENT
    global INPUT_COL
    global CDNU_COL
    global OUTPUT_COL
    global ERROR_COL

    rec_waitfor = re.compile("^Wait (for)?(at least)? ?(?P<Value>.*)(?P<TimeUnit>second|seconds|minute|minutes)",
                             re.IGNORECASE)

    # logging.info("In process_waitfor")

    wait = rec_waitfor.search(cell_val)

    if wait:
        logging.debug(f"{cell.row} Wait found in {cell_val}")
        try:
            waitval = wait.group('Value')
            timeunit = wait.group('TimeUnit')

            # convert ambigious values for 'few' and 'several' to be 10

            if 'few' in waitval or 'several' in waitval:
                waitval = 10

        except (NameError, AttributeError):
            logging.debug(f"{cell.row} Unable to determine wait time in {cell_val}")
            pass
        else:
            # Check to see if value is a number (e.g '2'), or english word of number (e.g 'two')
            # print ("Checking word values", cell_val)
            try:
                intval = int(waitval)
            except ValueError:
                # Number is not numeric. try to convert it into a number
                try:
                    intnum = w2n.word_to_num(waitval)
                except (NameError, AttributeError):
                    logging.debug(f"{cell.row} Wait time is non-numeric {cell_val}")
                    pass
                else:
                    if 'second' in timeunit.lower():
                        unit = 'S'
                    elif 'minute' in timeunit.lower():
                        unit = 'M'
                    elif ('millisec' in timeunit.lower()) or \
                            ('ms' in timeunit.lower()):  # Unlikely but added anyway
                        unit = 'MS'
                    else:
                        unit = 'UNKNOWN'

                    constructed_str = "WAIT" + SEPCH + str(intnum) + SEPCH + str(unit)
                    work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str

                    logging.debug(f"{cell.row} Wait for (non numeric) = {constructed_str}")

            # The number was numeric - so continue processing the rest
            else:
                if 'second' in timeunit.lower():
                    unit = 'S'
                elif 'minute' in timeunit.lower():
                    unit = 'M'
                elif ('millisec' in timeunit.lower()) or \
                        ('ms' in timeunit.lower()):  # Unlikely but added anyway
                    unit = 'MS'
                else:
                    unit = 'UNKNOWN'

                constructed_str = "WAIT" + SEPCH + str(intval) + SEPCH + str(unit)
                work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str

                logging.debug(f"{cell.row}  Wait for (numeric) = {constructed_str}")


def process_power_on_off_cdnu(cell_val, cell, work_sheet):
    """
    # Matches the power on/off/down setting

    :param cell_val:
    :param cell:
    :param work_sheet:
    """

    global SEPCH
    global COMMENT
    global INPUT_COL
    global CDNU_COL
    global OUTPUT_COL
    global ERROR_COL

    rec_cdnu = re.compile("Power\s(?P<State>on|OFF)\s(?P<CDNU>CDNU[12])", re.IGNORECASE)
    rec_cdnus = re.compile("Power\s(?P<State>on|OFF|down)\s(both|the)?\s?(?P<CDNU>CDNU'?[sS])", re.IGNORECASE)

    # logging.info("In process_power_on_off_cdnu")

    c1 = rec_cdnu.search(cell_val)

    if c1:
        try:
            state = c1.group('State')
            cdnu = c1.group('CDNU')
        except (NameError, AttributeError):
            logging.debug(f"{cell.row} Cant determine Power-State or which CDNU from {cell_val}")
            pass
        else:
            constructed_str = "RIG" + SEPCH + "SET" + SEPCH + str(cdnu) + SEPCH + str(state).upper()
            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            logging.debug(f"{cell.row} process_power_on_off_cdnu (match1) = {constructed_str}")

    # This part checks for both CDNUs, so if we get a match here, it will apply to Both CDNUs
    # rather than construct a string with CDNU1 then CDNU2, I have assumed CDNUS for both, so this
    # should be available

    c2 = rec_cdnus.search(cell_val)

    if c2:
        try:
            state = c2.group('State')
        except (NameError, AttributeError):
            logging.debug(f"{cell.row} Cant determine Power-State or which CDNU from {cell_val}")
            pass
        else:
            if state.lower() == "down":  # Turn 'down' to 'OFF'
                state = "OFF"

            constructed_str = "RIG" + SEPCH + "SET" + SEPCH + "CDNUS" + SEPCH + str(state).upper()
            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            logging.debug(f"{cell.row} process_power_on_off_cdnu (match2) = {constructed_str}")


def process_bus_analyser(cell_val, cell, work_sheet):
    """
    Processes the Bus Analyser commands.
    Any commands which set multiple values have been ignored.
    Bus Analyser: Set Squat switch to xxx
    Bus Analyser Set <Channel> to <Address> and also corresponding Word values e.g Word X to 16#FABC
    Bus Analyser Transmit <Channel> to <Address> and also corresponding Word values e.g Transmit Word X 16#FABC

    :param cell_val:
    :param cell:
    :param work_sheet:
    """

    global SEPCH
    global COMMENT
    global INPUT_COL
    global CDNU_COL
    global OUTPUT_COL
    global ERROR_COL
    r_num = 0  # row number

    rec_bus = re.compile("Bus Analyser:\s([Ss]et)\s(?P<Channel>[A-Z]{1,2}[0-9]{1,2})\s"
                         "(?P<Address>([A-Z]{1,2}[0-9]{1,2}))\s[Ww]ord\s"
                         "(?P<Word>\d{1,3})\sto\s(?P<Val>[0-9A-F]{1,8}#?[0-9A-F]{1,8})(?P<Last>.*)", re.IGNORECASE)

    rec_bus1 = re.compile("Bus Analyser: Set (?P<Channel>\w{3,5})\s(?P<Address>\w{3,5}) "
                          "words as follows:", re.IGNORECASE)

    rec_wd = re.compile("Word (?P<Word>\d{1,2}): (?P<WordLen>\d{1,2})#(?P<WordVal>([0-9A-F]){1,4}) "
                        "(?P<Last>.*)", re.IGNORECASE)

    rec_bus2 = re.compile("Bus Analyser: Transmit the following data for (?P<Channel>\w{3,5})\s(?P<Address>\w{3,5})")
    rec_ramp = re.compile("Word (?P<Word>\d{1,2}): Ramp up from (?P<WordLen1>\d{1,2})#(?P<WordVal1>([0-9A-F]){1,4}) "
                          "to (?P<WordLen2>\d{1,2})#(?P<WordVal2>([0-9A-F]){1,4}) in steps "
                          "of (?P<WordLen3>\d{1,2})#(?P<WordVal3>([0-9A-F]){1,4})(?P<Step>\d{1,2})(?P<Last>.*)",
                          re.IGNORECASE)

    ba = rec_bus.search(cell_val)
    ba1 = rec_bus1.search(cell_val)
    ba2 = rec_bus2.search(cell_val)

    if ba:
        try:
            ch = ba.group('Channel')
            add = ba.group('Address')
            wd = ba.group('Word')
            val = ba.group('Val')
            last = ba.group('Last')
        except (NameError, AttributeError):
            logging.debug(f"{cell.row} Cant determine Bus Analyser sub group from {cell_val}")
            # No match found - so just move on
            pass
        else:
            s_cdnu = work_sheet.cell(row=cell.row, column=CDNU_COL).value

            constructed_str = \
                str(s_cdnu) + SEPCH + \
                'SET' + SEPCH + \
                str(ch) + SEPCH + \
                str(add) + SEPCH + \
                str(wd) + SEPCH + \
                str(val) + SEPCH + \
                COMMENT + str(last)

            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            logging.debug(f"{cell.row} process_bus_analyser = {constructed_str}")

    if ba1:
        try:
            ch = ba1.group('Channel')
            add = ba1.group('Address')
        except (NameError, AttributeError):
            logging.debug(f"{cell.row} Cant determine Bus Analyser sub group 1from {cell_val}")
            # No match found - so just move on
        else:
            # Loop through the rest of the file, until we come to something not starting with Word

            while True:

                r_num = r_num + 1
                new_val = str(work_sheet.cell(row=cell.row + r_num, column=INPUT_COL).value)

                try:
                    w1 = rec_wd.search(new_val)
                except (NameError, AttributeError):
                    logging.debug(f"{cell.row} Breaking from Checking Word value {cell_val}")
                    return
                else:
                    if w1:
                        wd = w1.group('Word')
                        # wlen = w1.group('WordLen') # Future use - if 32/64 bit words are used, then this can be used
                        wval = w1.group('WordVal')
                        last = w1.group('Last')

                        constructed_str = \
                            '1553' + SEPCH + \
                            'SET' + SEPCH + \
                            str(ch) + SEPCH + \
                            str(add) + SEPCH + \
                            str(wd) + SEPCH + \
                            str(wval) + SEPCH +\
                            COMMENT + str(last)

                        work_sheet.cell(row=cell.row + r_num, column=OUTPUT_COL).value = constructed_str
                        # print(f"{cell.row + r_num} - {constructed_str}")
                        logging.debug(f"{cell.row} process_bus_analyser = {constructed_str}")
                    else:
                        break
    # Search for Bus Analyser: Transmit the following data for xxx
    if ba2:
        try:
            ch = ba2.group('Channel')
            add = ba2.group('Address')
        except (NameError, AttributeError):
            logging.debug(f"{cell.row} Cant determine Bus Analyser sub group 2 from {cell_val}")
            # No match found - so just move on
        else:
            # Loop through the rest of the file, until we come to something not starting with Word or Ramp
            while True:

                r_num = r_num + 1
                new_val = str(work_sheet.cell(row=cell.row + r_num, column=INPUT_COL).value)

                try:
                    w1 = rec_wd.search(new_val)
                    w2 = rec_ramp.search(new_val)
                except (NameError, AttributeError):
                    logging.debug(f"{cell.row} Breaking from Checking Word Ba2 new-value {cell_val}")
                else:
                    if w1:
                        # w_len1 = w1.group('WordLen') # Future use - if 32/64 bit words are used, then this can be used
                        wd = w1.group('Word')
                        val = w1.group('WordVal')
                        last = w1.group('Last')

                        constructed_str = \
                            '1553' + SEPCH + \
                            'SET' + SEPCH + \
                            str(ch) + SEPCH + \
                            str(add) + SEPCH + \
                            str(wd) + SEPCH + \
                            str(val) + SEPCH + \
                            COMMENT + str(last)

                        work_sheet.cell(row=cell.row + r_num, column=OUTPUT_COL).value = constructed_str
                        # print(f"BA2 {cell.row + r_num} - {constructed_str}")
                        logging.debug(f"{cell.row} BA2 {cell.row} process_bus_analyser = {constructed_str}")

                    elif w2:
                        # w_len1 = w1.group('WordLen1') # Future use - if 32/64 bit words are used
                        # w_len2 = w1.group('WordLen2') # Future use - if 32/64 bit words are used
                        # w_len3 = w1.group('WordLen3') # Future use - if 32/64 bit words are used
                        wd = w2.group('Word')
                        val1 = int(w2.group('WordVal1'), 16)   # Convert from Hex String to int
                        val2 = int(w2.group('WordVal2'), 16)
                        step = int(w2.group('Step'))
                        last = w2.group('Last')

                        constructed_str = ""

                        # need to loop around here for the step increment
                        for i in range(val1, val2, step):

                            constructed_str = constructed_str + \
                                '1553' + SEPCH + \
                                'SET' + SEPCH + \
                                str(ch) + SEPCH + \
                                str(add) + SEPCH + \
                                str(wd) + SEPCH + \
                                str(i) + SEPCH +\
                                COMMENT + str(last) + "\n"

                            # print(f"BA3{cell.row + r_num} - {constructed_str}")
                            logging.debug(f"{cell.row} BA3 {cell.row} process_bus_analyser = {constructed_str}")

                        work_sheet.cell(row=cell.row + r_num, column=OUTPUT_COL).value = constructed_str

                    else:
                        break


def process_test_rig(cell_val, cell, work_sheet):
    """
        Matches Test Rig: Set Squat swtich to xxx
        Group 1 is optional junk
        Group 2 is the switch
        Group 3 is the on/off state
    """

    global SEPCH
    global COMMENT
    global INPUT_COL
    global CDNU_COL
    global OUTPUT_COL
    global ERROR_COL

    set_to = re.search("Test Rig:\s[Ss]et( the)?\s(.*) to\s(.*)", cell_val)

    if set_to:
        try:
            switch_name = set_to.group(2)
            switch_state = set_to.group(3)
        except (NameError, AttributeError):
            logging.debug(f"{cell.row} Cant determine Process Test Rig sub group from {cell_val}")
            pass
        else:
            logging.debug(f"{cell.row} name = {switch_name} state =  {switch_state}")

            s_cdnu = work_sheet.cell(row=cell.row, column=CDNU_COL).value

            constructed_str = \
                str(s_cdnu) + SEPCH + \
                "RIG" + SEPCH + \
                "SET" + SEPCH + \
                str(switch_name) + SEPCH + \
                str(switch_state)

            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            logging.debug(f"{cell.row} process_test_rig = {constructed_str}")


def process_inspect(cell_val, cell, work_sheet):
    """
        Matches Inspect(n): LKn<something>### Inspect(n): <something3)
        Group 1 is LK key
        Group 2 is stuff  but now needed for upper/lower recognition
        Group 3 is junk between LK and the Comments
        group 4 is the values to set to which will be manipulated further
    """

    global SEPCH
    global COMMENT
    global INPUT_COL
    global CDNU_COL
    global OUTPUT_COL
    global ERROR_COL

    rec_srch_inspect1 = re.compile("(Inspect\(\d{1,3}\):\s+)(?P<LK>LK[0-9]) - (.*)###(.*):(?P<Set>.*)=(?P<To>.*)")
    srch_inspect1 = rec_srch_inspect1.search(cell_val)
    srch_inspect_comment = re.search(
            "Inspect\s?\(\d{1,3}\)\s?:.*(LK[0-9])\s(.*)(###\s?Inspect\s?\(\d{1,3}\)\s?:\s)(.*)", cell_val)

    if srch_inspect_comment:
        try:
            lk_str = srch_inspect_comment.group(1)
        except (NameError, AttributeError):
            logging.debug(f"{cell.row} Unable to get sub-groups for Inspect Comment Group 1 - {lk_str}")
            pass
        else:
            try:
                set_string = str(srch_inspect_comment.group(4))
            except (NameError, AttributeError):
                logging.debug(f"{cell.row} Unable to get sub-groups for Inspect Comment group 4 - {set_string}")
                pass
            else:
                try:
                    srch_val = re.search("(.*)\sis\s(.*)", set_string)
                    before_is = srch_val.group(1)
                    to_val = srch_val.group(2)
                    keyword = before_is.split()[-1]
                except (NameError, AttributeError):
                    logging.debug(f"{cell.row} Unable to get sub-groups for Inspect Comment  - {srch_val}")
                    pass
                else:

                    # default to upper line
                    line_num = 1
                    if (cell_val.lower()).find(" upper ") != -1:
                        line_num = 1

                    if (cell_val.lower()).find(" lower ") != -1:
                        line_num = 2

                    logging.debug(f"{cell.row} LK=[{lk_str}], set=[{set_string}], before_is = [{before_is}], to_val=[{to_val}], \
                            keyword = [{keyword}], Line = [{line_num}], {cell_val}")

                    s_cdnu = work_sheet.cell(row=cell.row, column=CDNU_COL).value

                    constructed_str = \
                        str(s_cdnu) + SEPCH + \
                        "INSPECT" + SEPCH + \
                        "DISPLAY" + SEPCH + \
                        str(lk_str) + SEPCH + \
                        str(line_num) + SEPCH + \
                        "EQUALTO" + SEPCH + \
                        str(to_val) + SEPCH + \
                        COMMENT + str(set_string)
                    work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str

    if srch_inspect1:
        try:
            lk = srch_inspect1.group('LK')
            set_val = srch_inspect1.group('Set').strip()
            to_val = srch_inspect1.group('To').strip()
        except (NameError, AttributeError):
            logging.debug("{cell.row} Unable to get sub-groups for Inspect1")
            pass
        else:
            s_cdnu = work_sheet.cell(row=cell.row, column=CDNU_COL).value

            constructed_str = \
                str(s_cdnu) + SEPCH + \
                "INSPECT" + SEPCH + \
                "DISPLAY" + SEPCH + \
                str(lk) + SEPCH + \
                str(set_val) + SEPCH + \
                "EQUALTO" + SEPCH + \
                str(to_val) + SEPCH

            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            work_sheet.cell(row=cell.row, column=ERROR_COL).value = 'ALERT! - Check Value Range'
            logging.debug(f"{cell.row} {constructed_str}\t\t,from {cell_val}")


def get_procedure_name(id_str: str, wrk_book: Workbook, procedure_file: str) -> str:
    """
    Scans through an Excel filename expecting two columns:
    The data must be in Sheet1
    Col A has the DOORS Id where the procedure is located. This is just the absolute id, with no prefixes
    Col B has the name of the DOORS procedure name.  The spaces are replaced by underscores prior to using the file.
    Both Col A and Col B must be populated, although some error checking does take place.
    Where the entry cannot be found, and ALERT text is returned, which can then be searched for in the converted file
    for easy modification
    """

    global ERROR_COL
    global PROCEDURE_COL

    # wrk_book = open_excel(procedure_file)
    wsheet = wrk_book.active
    found = False

    for cell in wsheet['A']:
        # print (cell.value)
        s = str(cell.value)

        if s.find(id_str) != -1:
            proc_name = wsheet.cell(row=cell.row, column=PROCEDURE_COL).value

            if proc_name is None:
                print(f"A Corresponding Procedure Name was not found for Id: {id_str} in {procedure_file}")
                return "ALERT! PROCEDURE NAME NOT FOUND"
            else:
                return proc_name

    if not found:
        print(f"No match found for {id_str} in {procedure_file}")
        return "ALERT! NO MATCH FOUND IN PROCEDURE FILE"


def process_keywords(wbk_test_script: Workbook, wbk_procedures: Workbook, xl_procedures: str):
    global SEPCH
    global COMMENT
    global INPUT_COL
    global CDNU_COL
    global OUTPUT_COL
    global ERROR_COL

    work_sheet = wbk_test_script.active                              # Select the active worksheet

    rec_keys = re.compile(
        "LK[0-9]|ALRT|COM|DATA|FPLN|NAV|SNSR|STR|TEST|WPT|FWD|BAK|BCK|CLR|ENT|"
        "BRT|DIM|LL_GRID|HDR|QUIT|PERF|HUMS|DF|IFF|IDM|TAC|BMN|GODIRECT|ON/OFF|"
        "-->|<--|<<-|(?P<MKFX>MARK\s*/*\s*FIX)|LBCK|LFWD|LCLR|LENT|LLK1|LLK2|LLK3|LLK4|LLK5|"
        "LRK1|LRK2|LRK3|LRK4|LRK5")

    rec_as_in = re.compile("[Aa]s in [Ss]ection|[Aa]s [Ss]ection|[Aa]s in ID")
    rec_id = re.compile("[iI][Dd]")

    for cell in work_sheet['B']:

        # cellval =  wsheet.cell(row = row, column = 2).value

        cellval = str(cell.value)

        if cellval is not None:  # Ignore blank lines

            re_as_in = rec_as_in.search(cellval)
            re_id = rec_id.search(cellval)
            re_inspect = re.search("Inspect", cellval)
            re_keys = rec_keys.search(cellval)

            # Get the CDNU allocation
            s_cdnu = work_sheet.cell(row=cell.row, column=CDNU_COL).value

            if re_inspect:
                # Ignore lines with Inspect lines for now
                pass
            elif re_as_in and re_id and not re_inspect:   # If it contains a "as in" & "id" it is probably a Procedure
                id_val = re.search("[Ii][Dd].(\d{1,7})", cellval)
                id_str = str(id_val.group(1))
                proc_name = get_procedure_name(id_str, wbk_procedures, xl_procedures)
                work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = s_cdnu + SEPCH + "PROC:" + proc_name

            elif re_keys:
                if s_cdnu is None:
                    work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = "ALERT! CDNU NOT DETERMINED"
                else:
                    # wsheet.cell(row=row, column=3).value = lastCDNU
                    # print('{0} cdnu={1} keyword [{2}], cell [{3}], last = [{4}]: sComment = {5}'.format(cell.row,
                    # s_cdnu, re_keys.group(), cellval, cellval[re_keys.end():], COMMENT))

                    # Process special cases
                    if re_keys.group('MKFX'):
                        # if re.search("MARK\s*/*\s*FIX", cellval):
                        s_construct = s_cdnu + SEPCH \
                                      + 'MARKFIX' \
                                      + SEPCH + COMMENT \
                                      + cellval[re_keys.end():]
                    else:
                        s_construct = s_cdnu + SEPCH \
                                     + re_keys.group()  \
                                     + SEPCH + COMMENT \
                                     + cellval[re_keys.end():]

                    work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = s_construct


def process_cdnu_allocation(wrk_book: Workbook):
    """
        Recognizes the current CDNU to operate on. it does this by going through the file once and
        looking for the CDNU allocation keywords e.g On CDNU1, or on Both CDNUs etc
        This is then used by several other processing functions
    """

    global CDNU_COL

    wsheet = wrk_book.active  # Select active worksheet

    last_cdnu = "CDNU1"

    print("Processing CDNU Allocations...")

    # for row in range (2, wsheet.max_row):
    for cell in wsheet['B']:

        # cellval =  wsheet.cell(row = row, column = 2).value

        cellval = str(cell.value)

        if cellval is not None:  # Ignore blank lines

            cdnu1_srch = re.search("^[oO]n\sCDNU.?1.", cellval)
            cdnu2_srch = re.search("^[oO]n\sCDNU.?2.", cellval)
            cdnus_srch = re.search("^[oO]n\s[Bb]oth\sCDNU.:", cellval)
            actions_srch = re.search("^Actions.*", cellval)

            # Add an '*' in the rows converted, as these will be removed in the final stage

            if cdnus_srch:
                last_cdnu = "CDNUS"
                logging.debug(f'{cell.row} CDNU Selection = {last_cdnu} from {cellval}')
                wsheet.cell(row=cell.row, column=CDNU_COL).value = '*'
            elif cdnu1_srch:
                last_cdnu = "CDNU1"
                logging.debug(f'{cell.row} CDNU Selection = {last_cdnu} from {cellval}')
                wsheet.cell(row=cell.row, column=CDNU_COL).value = '*'

            elif cdnu2_srch:
                last_cdnu = "CDNU2"
                logging.debug(f'{cell.row} CDNU Selection = {last_cdnu} from {cellval}')
                wsheet.cell(row=cell.row, column=CDNU_COL).value = '*'

            elif actions_srch:
                last_cdnu = "CDNU1"
                logging.debug(f'{cell.row} Selection = Resetting CDNU to {last_cdnu}')
            else:
                logging.debug(f'{cell.row} Selection = Using Default CDNU from {cellval}:')
                wsheet.cell(row=cell.row, column=CDNU_COL).value = last_cdnu

    print("Finished CDNU Allocations...")


def process_1553(cell_value, cell, work_sheet):
    """
    Processes the value of the argument cell_value against the various forms of 1553 Simulation commands
    The output is displayed (for the time being) and also modifies the appropriate work_sheet
    1553 Commands with multiple Set commands are ignored as there are too few to bother with at the moment
    """

    global SEPCH
    global COMMENT
    global OUTPUT_COL
    global ERROR_COL

    to_val = ""
    word_val = ""
    num_base = ""
    srch_to_grp = ""

    rec_srch_1553 = re.compile("1553 Simulator:")
    rec_srch_set = re.compile("[Ss]et")
    rec_srch_words = re.compile("[Ww]ords \d{1,2}")  # Multiple words in setting
    rec_srch_enable = re.compile(".*(1553 Simulat.*:)\s([Ee]nable)\s(.*)")
    rec_srch_disable = re.compile(".*(1553 Simulat.*:)\s([Dd]isable)\s(.*)")

    rec_srch_channel = re.compile("RT\d{1,2}")  # Channel starts with RTnn (n= 0-9)
    rec_srch_address = re.compile("SA\d{1,3}")  # Address starts with STnnn (n = 0-9)
    rec_srch_word = re.compile("([Ww]ord|[Ww]rd) \d{1,2}")  # Word identifier starts with Word nn
    rec_srch_to = re.compile("([wW]ord|[Bb]it)? to\s([Hh]ex|[Dd]ec|[Bb]in)?([0-9A-F ]*)")  # Get 'to'rec_

    rec_srch_to_grp = re.compile("[0-9A-F ]*")
    rec_srch_base = re.compile("[Hh]ex|[Dd]ec")
    rec_srch_bracket = re.compile("\(.+\)")  # additional information in brackets

    srch_1553 = rec_srch_1553.search(cell_value)
    srch_set = rec_srch_set.search(cell_value)
    srch_words = rec_srch_words.search(cell_value)
    srch_enable = rec_srch_enable.search(cell_value)
    srch_disable = rec_srch_disable.search(cell_value)

    srch_channel = rec_srch_channel.search(cell_value)  # Channel starts with RTnn (n= 0-9)
    srch_address = rec_srch_address.search(cell_value)  # Address starts with STnnn (n = 0-9)
    srch_word = rec_srch_word.search(cell_value)  # Word identifier starts with Word nn
    srch_to = rec_srch_to.search(cell_value)  # Get

    if srch_1553 and srch_set:

        if srch_to is not None:
            srch_to_grp = re.search("\d{1,5}", srch_to.group())

        srch_bracket = rec_srch_bracket.search(cell_value)                # additional information in brackets

        if srch_channel:                                                # Get Channel match
            channel_val = srch_channel.group()                          # Get Actual Channel number
        else:
            channel_val = "No CH"

        if srch_address:                                                # Get Address match
            address_val = srch_address.group()                          # Get Address number
        else:
            address_val = "No ADD"

        if srch_word:                                                   # For single word settings
            srch_wrd_grp = re.search("\d{1,2}", srch_word.group())      # Match the word number
            if srch_wrd_grp:
                word_val = srch_wrd_grp.group()                         # Get the actual number

        if srch_to:
            srch_to_grp = rec_srch_to_grp.search(srch_to.group(3))
            srch_base = rec_srch_base.search(cell_value)

            if srch_to_grp:
                to_val = srch_to_grp.group().rstrip()

            if srch_base:
                if srch_base.group() == "hex".lower():
                    num_base = 'H'                                          # hex value
                elif srch_base.group() == "dec".lower():
                    num_base = 'D'
                elif srch_base.group() == "bin".lower():
                    num_base = 'B'
                else:
                    num_base = ''                                           # => most of the time, could default to Hex

        if srch_bracket:
            comment_str = srch_bracket .group()
        else:
            comment_str = ""

        if srch_channel and srch_address and srch_word and srch_word and srch_to and srch_to_grp:
            constructed_str = "1553:SET" + SEPCH + \
                              channel_val + SEPCH + \
                              address_val + SEPCH + \
                              word_val + SEPCH + \
                              to_val + SEPCH + \
                              num_base + SEPCH +\
                              COMMENT + comment_str
            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str

            logging.debug(f"{cell.row} 1553(ALL): {constructed_str}, [{cell_value}]")
        # No Channel
        elif srch_address and srch_word and srch_word and srch_to and srch_to_grp and channel_val == "No CH":
            constructed_str = "1553:SET" + SEPCH + \
                              address_val + SEPCH + \
                              word_val + SEPCH + \
                              to_val + SEPCH + \
                              num_base + SEPCH +\
                              COMMENT + comment_str
            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            work_sheet.cell(row=cell.row, column=ERROR_COL).value = "No 1553 CH"

            logging.debug(f"{cell.row} 1553(NO CH): {constructed_str}, [{cell_value}] ")
        elif srch_words:  # Dont process multiple word settings - too few and complicated
            logging.debug(f"{cell.row} 1553: Found Multiple word settings - Ignoring, [{cell_value}]")
        else:
            constructed_str = "1553:SET" + SEPCH + \
                              channel_val + SEPCH + \
                              address_val + SEPCH + \
                              word_val + SEPCH + \
                              to_val + SEPCH + \
                              num_base + SEPCH +\
                              COMMENT + comment_str
            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            work_sheet.cell(row=cell.row, column=ERROR_COL).value = "ALERT!! - PLS CHECK"
            logging.debug(f"{cell.row} 1553(Other Issue!!): {constructed_str}, [{cell_value}] ")

    if srch_disable:
        srch_disable_channel = srch_disable.group(3)

        if srch_disable_channel:
            constructed_str = "1553:SET" + SEPCH + srch_disable_channel + SEPCH + '0'
            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            logging.debug(f"{cell.row} DISABLING {constructed_str}")
        else:
            srch_disable_channel = "No 1553 Channel"
            constructed_str = "1553:SET" + SEPCH + srch_disable_channel + SEPCH + '0'
            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            work_sheet.cell(row=cell.row, column=ERROR_COL).value = "NO DISABLE CH"
            logging.debug(f"{cell.row} NO DISABLE CH {constructed_str}")

    if srch_enable:
        try:
            srch_enable_channel = srch_enable.group(3)
            constructed_str = "1553:SET" + SEPCH + srch_enable_channel
            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            logging.debug(f"{cell.row} Enabling {constructed_str}")
        except(NameError, AttributeError):
            srch_enable_channel = "No 1553 Channel"
            constructed_str = "1553:SET" + SEPCH + srch_enable_channel
            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = constructed_str
            work_sheet.cell(row=cell.row, column=ERROR_COL).value = "NO ENABLE CH"
            logging.debug(f"{cell.row} NO ENABLE CH {constructed_str}")


def new_process_keywords(cell_value, cell, work_sheet, wbk_procedures: Workbook, xl_procedures: str):
    """
        Processes the main CDNU Key keywords, e.g DATA, FPLN, LK1 etc
        There is a special consideration for Mark Fix, as the output required doesnt match the input form
    """

    global SEPCH
    global COMMENT
    global INPUT_COL
    global CDNU_COL
    global OUTPUT_COL
    global ERROR_COL

    rec_keys = re.compile(
        "LK[0-9]|ALRT|COM|DATA|FPLN|NAV|SNSR|STR|TEST|WPT|FWD|BAK|BCK|CLR|ENT|"
        "BRT|DIM|LL_GRID|HDR|QUIT|PERF|HUMS|DF|IFF|IDM|TAC|BMN|GODIRECT|ON/OFF|"
        "-->|<--|<<-|(?P<MKFX>MARK\s*/*\s*FIX)|LBCK|LFWD|LCLR|LENT|LLK1|LLK2|LLK3|LLK4|LLK5|"
        "LRK1|LRK2|LRK3|LRK4|LRK5")

    rec_as_in = re.compile("[Aa]s in [Ss]ection|[Aa]s [Ss]ection|[Aa]s in ID")
    rec_id = re.compile("[iI][Dd]")

    if cell_value is not None:  # Ignore blank lines

        re_as_in = rec_as_in.search(cell_value)
        re_id = rec_id.search(cell_value)
        re_inspect = re.search("Inspect", cell_value)
        re_keys = rec_keys.search(cell_value)

        # Get the CDNU allocation
        s_cdnu = work_sheet.cell(row=cell.row, column=CDNU_COL).value

        if re_as_in and re_id and not re_inspect:  # If it contains a "as in" & "id" it is probably a Procedure
            id_val = re.search("[Ii][Dd].(\d{1,7})", cell_value)
            id_str = str(id_val.group(1))
            proc_name = get_procedure_name(id_str, wbk_procedures, xl_procedures)
            work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = s_cdnu + SEPCH + "PROC:" + proc_name

        elif re_keys:
            if s_cdnu is None:
                work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = "ALERT! CDNU NOT DETERMINED"
            else:
                # Process special cases
                if re_keys.group('MKFX'):
                    s_construct = s_cdnu + SEPCH \
                                  + 'MARKFIX' \
                                  + SEPCH + COMMENT \
                                  + cell_value[re_keys.end():]
                else:
                    s_construct = s_cdnu + SEPCH \
                                  + re_keys.group() \
                                  + SEPCH + COMMENT \
                                  + cell_value[re_keys.end():]

                work_sheet.cell(row=cell.row, column=OUTPUT_COL).value = s_construct


# # ############################################# Main #####################################################
#
# # wbook = load_workbook("e:/temp/py/x1.xlsx") ### = this works
#
#
#
# excel_script_file = "C:/temp/newragu7-newcode.xlsx"
# excel_procedure_file = "H:/Dev/Python/TestParser/Procedures.xlsx"
# logf# ile = 'H:/Dev/Python/TestParser/TestScript.log'
#
# # Configure the Logger - change to using a config file later
# # TODO - use config file at some point
#
# logging.basicConfig(handlers=[logging.FileHandler(logfile, 'w', 'utf-8')],
#                     level=logging.DEBUG,
#                     format='%(asctime)s - %(levelname)-10s - %(message)s',
#                     datefmt='%d-%b-%y %H:%M:%S')
#
#
# # Open the Excel file_names
# wb_script = open_excel(excel_script_file)
# wb_procedures = open_excel(excel_procedure_file)
#
# # Process
# process_cdnu_allocation(wb_script)                                  # figure out the CDNU for each command
# # process_keywords(wb_script, wb_procedures, excel_procedure_file)  # process the procedure Names
#
# # The following is the preferred method for executing functions.
#
# worksheet = wb_script.active                                        # Select active worksheet
#
# # for row in range (2, wsheet.max_row):
# for cell in worksheet['B']:
#     # cellval =  wsheet.cell(row = row, column = 2).value
#     cell_val = str(cell.value)
#
#     process_inspect(cell_val, cell, worksheet)
#     process_test_rig(cell_val, cell, worksheet)
#     process_bus_analyser(cell_val, cell, worksheet)
#     process_power_on_off_cdnu(cell_val, cell, worksheet)
#     process_waitfor(cell_val, cell, worksheet)
#     process_arinc(cell_val, cell, worksheet)
#     process_1553(cell_val, cell, worksheet)
#     new_process_keywords(cell_val, cell, worksheet, wb_procedures, excel_procedure_file)
#
# # Close Filenames
# close_excel(wb_script, excel_script_file)
# close_excel(wb_procedures, excel_procedure_file)
#
# print(f"\n\n Please check the logging information in {logfile}")


def showusage(myname: str):
    """
        When running the script in command line, the options which can be provided are shown here
    """

    print(f"\nUsage:\n\tpython.exe {myname} [-i | --infile] <inputfile> "
          f"[-o | --outfile] <outputfile> [-l | --logfile] <logfile>\n"
          "\t-i or --infile   is the Input script file (expected as Excel .xlsx)\n"
          "\t-p or --procfile is the Procedures index file (as Excel .xlsx)\n"
          "\t-l or --logfile  is the Logfle for Debug purposes\n"
          "\tThe results are placed into the inputfile, which must be closed when running this process")


def process_command_line(argv):
    """
        The script can be run either in a command line format, for batch processing, or via the GUI
        When running in a commandline, this parses the input options for selection of the filenames
        and then runs the processing engine
    """

    excel_script_file = ''
    excel_procedure_file = ''
    logfile = ''

    try:
        opts, args = getopt.getopt(argv, "hi:p:l:", ["infile=", "procfile=", "logfile="])

    except getopt.GetoptError as e:
        print("\n\n", str(e))
        showusage(sys.argv[0])
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            showusage(sys.argv[0])
            sys.exit()

        elif opt in ("-i", "--infile"):
            excel_script_file = arg

        elif opt in ("-p", "--procfile"):
            excel_procedure_file = arg

        elif opt in ("-l", "--logfile"):
            logfile = arg


    if excel_script_file == '' or excel_procedure_file == '' or logfile == '':
        print ("Must supply all three inputs")
        showusage(sys.argv[0])
    else:
        print("\nProcessing script using following")
        print(f"Input file      = {excel_script_file}")
        print(f"Procedures file = {excel_procedure_file}")
        print(f"logfile file    = {logfile}")

        run_processing_engine(excel_script_file, excel_procedure_file, logfile, False)

        print(f"\n\nLogging information captured in {logfile}")


def run_processing_engine(script_file: str, procedure_file: str, logfile: str, with_gui: bool):

    # Setup the Logfile
    logging.basicConfig(handlers=[ logging.FileHandler(logfile, 'w', 'utf-8')],
                        level=logging.DEBUG,
                        format='%(asctime)s - %(levelname)-10s - %(message)s',
                        datefmt='%d-%b-%y %H:%M:%S')

    # Open the Excel file_names
    wb_script = open_excel(script_file)
    wb_procedures = open_excel(procedure_file)

    # Process

    if with_gui:
        app.t_out.insert('end', "Processing CDNU Allocations..\n")
        app.update_idletasks()

    process_cdnu_allocation(wb_script)                                  # figure out the CDNU for each command
    worksheet = wb_script.active                                        # Select active worksheet

    if with_gui:
        app.t_out.insert('end', "Processing Script...\n")
        app.update_idletasks()

    # for row in range (2, wsheet.max_row):
    for cell in worksheet['B']:
        # cellval =  wsheet.cell(row = row, column = 2).value
        cell_val = str(cell.value)

        if with_gui:
            cell_lf = cell_val + '\n'
            app.t_out.insert('end', cell_lf)
            app.t_out.see('end')
            app.update_idletasks()

        process_inspect(cell_val, cell, worksheet)
        process_test_rig(cell_val, cell, worksheet)
        process_bus_analyser(cell_val, cell, worksheet)
        process_power_on_off_cdnu(cell_val, cell, worksheet)
        process_waitfor(cell_val, cell, worksheet)
        process_arinc(cell_val, cell, worksheet)
        process_1553(cell_val, cell, worksheet)
        new_process_keywords(cell_val, cell, worksheet, wb_procedures, procedure_file)

    # format the output column(s) as desired
    for r in range(2, worksheet.max_row):
        worksheet.cell(row=r, column=CDNU_COL).font = Font(name='Calibri', size=10)
        worksheet.cell(row=r, column=OUTPUT_COL).font = Font(name='Calibri', size=10)
        worksheet.cell(row=r, column=ERROR_COL).font = Font(name='Calibri', size=10, color = colors.RED )


    # Close Filenames
    close_excel(wb_script, script_file)
    close_excel(wb_procedures, procedure_file)


class Window(Frame):

    help_text = """Use the buttons below to select the script file to process. 
    This file MUST be preprocessed by Labview so that the Inspect commands are aligned within the
    Test Steps as well as having each command on a separate row, the row numbers are processed 
    to align with ONE DOORS identifier per test. This MUST also be in the .XLSX excel file format.

    The procedures file MUST be in the .XLSX excel format
    It expects the Identifier in Column 1, and the Procedure name in Column 2
    The procedure names should not have any spaces 

    The Logfile captures more detailed information about the process for any debug purposes

    This window will show the progress, once the Process button is pressed.
    Note: Pressing the file selection butons below take a few seconds to launch
    """

    def __init__(self, master=None):
        TROW = 1
        SROW = 2
        PROW = 4
        LROW = 5
        BTNROW = 6

        COL = 2
        BTNCOL = 2

        # parameters that you want to send through the Frame class.
        Frame.__init__(self, master)

        # reference to the master widget, which is the tk window
        self.master = master
        self.master.title("Wildcat Script Parser v1.0")

        # Define Menu

        menu = Menu(self.master)  # creating a menu instance
        self.master.config(menu=menu)

        file = Menu(menu)  # create the file Menu
        file.add_command(label="Exit", command=self.menu_exit)  # adds a Exit to the menu option
        menu.add_cascade(label="File", menu=file)  # bind the function file to Menu "File" Label

        # Define Form

        # Define the Labels

        self.l1 = Label(self.master, text="Test Script")
        self.l2 = Label(self.master, text="Procedures File")
        self.l3 = Label(self.master, text="LogFile")

        self.l1.grid(row=SROW)
        self.l2.grid(row=PROW)
        self.l3.grid(row=LROW)

        # Define the Widgets

        self.t_out = scrolledtext.ScrolledText (self.master, height=15, width=100, fg='grey')
        self.t_scr = Text(self.master, width=100, height=1, font=('Ariel', 10))
        self.t_proc = Text(self.master, height=1, width=100, font=('Ariel', 10))
        self.t_log = Text(self.master, height=1, width=100, font=('Ariel', 10))

        # Define the positioning
        self.t_out.grid(row=TROW, column=COL)
        self.t_scr.grid(row=SROW, column=COL, sticky=W)
        self.t_proc.grid(row=PROW, column=COL, sticky=W)
        self.t_log.grid(row=LROW, column=COL, sticky=W)
        # self.t4.place (x=50,y=320)
        # l1.place(x=50,y=300)

        # Assign the widgets to procedures
        Button(self.master, text='Script File',
               command=self.get_script_file, width=12).grid(row=SROW, column=BTNCOL, sticky=E, pady=4)
        Button(self.master, text='Procedure File',
               command=self.get_procedure_file, width=12).grid(row=PROW, column=BTNCOL, sticky=E, pady=4)
        Button(self.master, text='Logfile',
               command=self.get_logfile, width=12).grid(row=LROW, column=BTNCOL, sticky=E, pady=4)
        Button(self.master, text='Process Script',
               command=self.process_script, width=12).grid(row=BTNROW, column=BTNCOL, sticky=E, padx=100)

        Button(self.master, text='Exit', command=self.menu_exit).place(x=85, y=350)

        self.t_out.delete('1.0', 'end')
        self.t_out.insert('end', self.help_text)

    def get_script_file(self):
        script_filename = filedialog.askopenfilename(initialdir="/", title="Select Script file",
                                                     filetypes=(("Excel Files", "*.xlsx"), ("all files", "*.*")))
        self.t_scr.delete('1.0', 'end')
        self.t_scr.insert('end', script_filename)

    def get_procedure_file(self):
        proc_filename = filedialog.askopenfilename(initialdir="/", title="Select Procedures file",
                                                   filetypes=(("Excel Files", "*.xlsx"), ("all files", "*.*")))
        self.t_proc.delete('1.0', 'end')
        self.t_proc.insert('end', proc_filename)

    def get_logfile(self):
        log_filename = filedialog.askopenfilename(initialdir="/", title="Select Logfile",
                                                  filetypes=(("Log file", "*.txt"), ("all files", "*.*")))
        self.t_log.delete('1.0', 'end')
        self.t_log.insert('end', log_filename)

    def menu_exit(self):
        exit()

    def process_script(self):

        self.t_out.delete('1.0', 'end')

        script_file = self.t_scr.get('1.0', 'end-1c')           # "end - 1c" removes \n from text
        procedure_file = self.t_proc.get('1.0', 'end-1c')
        logfile = self.t_log.get('1.0', 'end-1c')

        self.t_out.insert('end', 'Using Script file: {}\n'.format(script_file))
        self.t_out.insert('end', 'Using Procedure file: {}\n'.format(procedure_file))
        self.t_out.insert('end', 'Using Logfile: {}\n'.format(logfile))

        run_processing_engine(script_file, procedure_file, logfile, True)

        self.t_out.insert('end', "\nFinished")
        self.t_out.see('end')



# #########################################################################
# # MAIN
# #########################################################################

if __name__ == "__main__":

    if len(sys.argv) > 1:
        process_command_line(sys.argv[1:])
    else:
        root = Tk()
        root.geometry("950x400")
        app = Window(root)                                         # creation of an instance
        root.mainloop()                                            # mainloop
