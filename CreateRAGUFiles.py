"""
A. Moghul - May 2019

Input file
~~~~~~~~~~
The .xlsx  RAGU format translated excel spreadsheet.  This is the output from TranslateDOORSScript.py

Output file(s)
~~~~~~~~~~~~~~
This script will take the input file (.xlsx) and generate multiple files in the format
DOORSMMODULE@ID.xlsx.

Purpose:
~~~~~~~~
These files will then be used as input by a DOS/DXL program, which will automatically insert the files
as objects into the DOORS module.

Notes
~~~~~~
IMPORTANT:  The script relies on the format of the Excel file being as expected. Any deviation from the expected
format will lead to files being generated in an unexpected manner, or cause the script to fail.

"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import logging
import sys
import getopt
import os

# GLOBAL Definitions

ID_COL = 1
INPUT_COL = 2
OUTPUT_COL = 4
CH='@'


def open_excel(xl_filename: str) -> Workbook:
    """
    Opens an Excel file for processing, given the pathname
    :param xl_filename:
    :return:  Workbook
    """
    logging.info("open_excel")

    try:
        wbook = load_workbook(xl_filename)  # Load file
        logging.info(f"Opened file {xl_filename}")
        return wbook
    except FileNotFoundError:
        print(f"File {xl_filename} not found, Exiting...")
        exit(-1)
    except OSError:
        print(f"{xl_filename} Error Saving file... Exiting...")
        exit(-1)
    except PermissionError:
        print(f"Error reading file {xl_filename} ... Exiting...")
        exit(-1)
    except Exception as ex:
        print(f'Error when opening {xl_filename} :', (str(ex)))
        exit(-1)


def close_excel(wbook: Workbook, xl_filename: str):
    """
    Closes the excel file

    :param wbook:
    :param xl_filename:
    :return:
    """
    logging.info("close_excel")

    try:
        wbook.save(xl_filename)
        wbook.close()
        logging.info(f"Closed file {xl_filename}")
    except FileNotFoundError:
        print(f"File {xl_filename} not found when Closing, Exiting...")
        exit(-1)
    except OSError:
        print(f"{xl_filename} Error Saving file... Exiting...")
        exit(-1)
    except PermissionError:
        print(f"{xl_filename} Error Saving file... Exiting...")
        exit(-1)
    except Exception as ex:
        print(f'Error when Closing {xl_filename} :', (str(ex)))
        exit(-1)


def showusage(myname: str):
    """
        When running the script in command line, the options which can be provided are shown here
    """

    print(f"\nUsage:\n\tpython.exe {myname} "
          f"[-i | --infile] <inputfile> "
          f"[-o | --outfolder] <outputfolder> "
          f"[-l | --logfile] <logfile>\n"
          "\t-i or --infile    is the Input script file (expected as Excel .xlsx)\n"
          "\t-o or --outfolder is the Output folder for generated files\n"    
          "\t-l or --logfile   is the Logfle for Debug purposes\n")


def process_command_line(argv):
    """
        The script can be run either in a command line format or directly from a python command shell
        The GUI facilities have been removed as it its not expected this will be run using a front end

        Input file MUST be a .xlsx formatted excel file, which was generated using translateDOORSscript.py
        The Output is a folder, where all the excel files will be created
            These files will be used by another DOS/DXL script to insert the files into DOOORS objects
        The Logfile is merely for debug purposes

    """

    excel_script_file = ''
    output_directory = ''
    logfile = ''

    try:
        opts, args = getopt.getopt(argv, "hi:o:l:", ["infile=", "output=", "logfile="])

    except getopt.GetoptError as e:
        print("\n\n", str(e))
        showusage(sys.argv[0])
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            showusage(sys.argv[0])
            sys.exit()

        elif opt in ("-i", "--infile"):
            excel_script_file = arg

        elif opt in ("-o", "--output"):
            output_directory = arg

        elif opt in ("-l", "--logfile"):
            logfile = arg

    if excel_script_file == '' or output_directory == '' or logfile == '':
        print("Please supply ALL inputs")
        showusage(sys.argv[0])
    else:
        print(f"excel_script_file = {excel_script_file}")
        print(f"output_directory = {output_directory}")
        print(f"logfile = {logfile}")
        print("Processing...")

        generate_RAGU_files(excel_script_file, output_directory, logfile)

        print(f"Finished\nLogging information captured in {logfile}")


def generate_RAGU_files(script_file: str, output_folder: str, logfile: str):

    cell_wb = None

    # Setup the Logfile
    logging.basicConfig(handlers=[ logging.FileHandler(logfile, 'w', 'utf-8')],
                        level=logging.DEBUG,
                        format='%(asctime)s - %(levelname)-8s - %(message)s',
                        datefmt='%d-%b-%y %H:%M:%S')

    if not os.path.exists(output_folder):                     # Check output directory exists before going too far
        try:
            os.mkdir(output_folder)
        except Exception as e:
            print("Unable to create diretory..", str(e))
            exit(2)

    wb_script = open_excel(script_file)                        # Open the Excel file_names
    worksheet = wb_script.active                               # Select active worksheet
    old_id_raw = worksheet.cell(row=1, column=ID_COL).value    # initialise
    old_id_list = old_id_raw.rsplit('/', 1)
    old_id = old_id_list.pop()
    cell_counter = 1

    for cell in worksheet['B']:

        cell_id = worksheet.cell(row=cell.row, column=ID_COL).value
        cell_action = worksheet.cell(row=cell.row, column=OUTPUT_COL).value

        stripped_list = cell_id.rsplit('/', 1)                 # strip out the doors module id
        real_id = stripped_list.pop()
        stripped_module = stripped_list.pop()

        if (stripped_module.find('ID : ')) == 0:
            print("Input file does not conform to expected format. There must be no ID prefix")
            sys.exit(2)

        module_str = stripped_module.replace('/', '_')

        if old_id == real_id:
            if cell_wb:
                cell_counter = cell_counter + 1
                pass
            else:
                cell_wb = Workbook()
                cell_ws = cell_wb.active
                cell_ws.cell(1, 1, "Actions")
                cell_ws.column_dimensions['A'].width = 150

            cell_ws.cell(cell_counter, 1).font = Font(size=10)
            cell_ws.cell(cell_counter, 1, cell_action)
            cell_ws.cell(cell_counter, 1).alignment = Alignment(horizontal='left')
            logging.debug(f"{real_id} - {cell_action}")

        else:
            full_pathname = output_folder + '/' + module_str + CH + old_id + CH + '.xlsx'
            logging.debug(f"Full Pathname = {full_pathname}")

            try:
                cell_wb.save(full_pathname)
                cell_wb.close()
            except Exception as e:
                print("Unable to save files in output directory. Aborting...", str(e))
                exit(2)
            else:
                cell_wb = None
                cell_counter = 2

        old_id = real_id

    # Close Filenames
    close_excel(wb_script, script_file)


# #########################################################################
# # MAIN
# #########################################################################

if __name__ == "__main__":

    if len(sys.argv) > 1:
        process_command_line(sys.argv[1:])
    else:
        showusage(sys.argv[0])
